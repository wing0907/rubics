from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import datetime

# 1) 입력 파일 경로
INPUT_XLSX = Path(r"D:\Users\Qube\0. 데이터\데이터분석_최종\데이터분석_큐브_최종_장우진.xlsx")

# 2) 출력 폴더 (권한 문제 방지용: 쓰기 가능한 곳 추천)
OUTPUT_DIR = Path(r"D:\Users\장우진\dev26\qube_out_readable")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# 3) 시트 설명 (탭명에 붙일 요약)
sheet_desc = {
    "summary_report": "전체 데이터 요약",
    "overall_images": "전체 이미지 요약",
    "daily_questions_qd": "일자별 통계 question_data",
    "daily_questions_ql": "일자별 통계 question_list",
    "subject_stats": "과목별 통계",
    "subject_top200": "과목 상위200",
    "subject_rank_top200": "과목 랭킹 상위200",
    "subject_rank_bottom200": "과목 랭킹 하위200",
    "daily_rank_top500": "일자 랭킹 상위",
    "daily_rank_bottom500": "일자 랭킹 하위",
    "class_stats": "class별 통계",
    "class_top200": "class 상위200",
    "image_per_question_dist": "질문당 이미지 분포",
    "img_dist_with_cum": "질문당 이미지 분포 누적",
    "decoded_index": "decoded 파일 인덱스",
    "meta": "메타정보",
    "missing_before": "결측 점검 전",
    "missing_after": "결측 점검 후",
}

# 4) 컬럼 설명 (헤더 셀에 같이 붙일 내용)
col_desc_common = {
    "rows": "해당 그룹 질문 건수",
    "valid_qid": "질문 ID(QM_QST_NO) 정상 건수",
    "valid_qid_rows": "질문 ID(QM_QST_NO) 정상 행 수",
    "date": "일자",

    "DomName": "과목 대분류",
    "SubName": "과목 소분류",
    "class_value": "반 또는 클래스 구분값",

    "image_questions": "이미지 포함 질문 건수",
    "total_images": "이미지 총 개수 합계",
    "avg_images_per_question": "질문 1건당 평균 이미지 수",
    "avg_msgs": "질문 1건당 평균 메시지 수",
    "image_rate": "이미지 포함 질문 비율(0~1)",
    "image_rate_pct": "이미지 포함 질문 비율(퍼센트)",
    "image_question_rate_pct": "이미지 포함 질문 비율(퍼센트)",

    "p50_first": "첫 응답 지연 t_first_response_s p50",
    "p90_first": "첫 응답 지연 t_first_response_s p90",
    "p95_first": "첫 응답 지연 t_first_response_s p95",
    "p50_done": "완료 지연 t_done_s p50",
    "p90_done": "완료 지연 t_done_s p90",
    "p95_done": "완료 지연 t_done_s p95",

    "images_per_question": "질문당 이미지 수(0,1,2..)",
    "questions": "해당 구간 질문 건수",
    "rate_pct": "해당 구간 비율(퍼센트)",
    "cum_questions": "누적 질문 건수",
    "cum_rate_pct": "누적 비율(퍼센트)",

    "generated_at": "리포트 생성 시각",
    "na_cnt": "NaN 개수",
    "na_rate_pct": "NaN 비율(퍼센트)",
    "blank_like_cnt_excluding_na": "NaN 제외 빈값 유사 케이스 개수",
    "blank_like_rate_pct_excluding_na": "NaN 제외 빈값 유사 비율(퍼센트)",

    "t_first_response_s": "첫 응답까지 걸린 시간(초)",
    "t_done_s": "완료까지 걸린 시간(초)",
}

def normalize_sheet_name(name: str) -> str:
    # 이미 시트명(설명) 형태면 기본 이름만 추출
    if "(" in name and name.endswith(")"):
        return name[:name.rfind("(")].strip()
    return name.strip()

def base_col_name(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    # 이미 컬럼명 + 설명 형태면 앞부분만
    if "\n" in s:
        return s.split("\n", 1)[0].strip()
    if " (" in s and s.endswith(")"):
        return s.split(" (", 1)[0].strip()
    return s

def infer_col_desc(col: str) -> str:
    c = col.strip()
    if c.startswith(("p50_", "p90_", "p95_")):
        q = c.split("_", 1)[0]
        rest = c.split("_", 1)[1] if "_" in c else ""
        if rest == "first":
            return f"첫 응답 지연 t_first_response_s {q}"
        if rest == "done":
            return f"완료 지연 t_done_s {q}"
        return f"{rest} 지표 {q}"
    if c.endswith("_rate"):
        return f"{c} 비율(0~1)"
    if c.endswith("_pct"):
        return f"{c} 비율(퍼센트)"
    if c.endswith("_cnt"):
        return f"{c} 개수"
    if c.startswith("avg_"):
        return f"{c} 평균 지표"
    return ""

def make_unique_sheet_title(wb, desired: str) -> str:
    # Excel 시트명 31자 제한 + 중복 방지
    desired = desired[:31]
    if desired not in wb.sheetnames:
        return desired

    base = desired[:28]  # 뒤에 _2 같은 suffix 붙일 공간
    i = 2
    while True:
        cand = f"{base}_{i}"
        cand = cand[:31]
        if cand not in wb.sheetnames:
            return cand
        i += 1

def safe_sheet_title(base: str, desc: str) -> str:
    base = base.strip()
    desc = desc.replace(" ", "").strip()
    if not desc:
        return base[:31]
    new = f"{base}({desc})"
    if len(new) <= 31:
        return new
    # 너무 길면 desc를 줄임
    max_desc_len = 31 - len(base) - 2
    if max_desc_len <= 0:
        return base[:31]
    return f"{base}({desc[:max_desc_len]})"

# 실행
wb = load_workbook(INPUT_XLSX)

# data_dictionary 시트 제거 (있으면)
if "data_dictionary" in wb.sheetnames:
    wb.remove(wb["data_dictionary"])

# 헤더 스타일
header_align = Alignment(wrap_text=True, vertical="center", horizontal="center")

# 시트별 작업
original_titles = [ws.title for ws in wb.worksheets]

for ws in wb.worksheets:
    orig_title = ws.title
    key = normalize_sheet_name(orig_title)

    # 1) 시트 탭명 변경
    desc = sheet_desc.get(key, "")
    desired_title = safe_sheet_title(key, desc) if desc else key[:31]
    desired_title = make_unique_sheet_title(wb, desired_title)
    ws.title = desired_title

# 탭명 변경 후, 다시 순회하면서 헤더 수정
for ws in wb.worksheets:
    # 2) 헤더(1행) 수정
    # 1행이 헤더가 아닌 시트는 그대로 두고 싶으면 조건을 더 걸 수 있음
    row1 = list(ws[1])
    if not row1:
        continue

    any_header = False
    for cell in row1:
        base = base_col_name(cell.value)
        if not base:
            continue

        any_header = True

        desc = col_desc_common.get(base, "")
        if not desc:
            desc = infer_col_desc(base)

        if desc:
            # 컬럼명 + 줄바꿈 + 설명 형태로 표시 (가독성 좋음)
            cell.value = f"{base}\n{desc}"
        else:
            cell.value = base

        cell.alignment = header_align

    # 헤더가 있는 시트만 높이 조정 + 고정
    if any_header:
        ws.row_dimensions[1].height = 34
        # 이미 freeze_panes가 있으면 유지, 없으면 헤더 고정
        if ws.freeze_panes is None:
            ws.freeze_panes = "A2"

# 저장
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
out_path = OUTPUT_DIR / f"{INPUT_XLSX.stem}_readable_{ts}.xlsx"
wb.save(out_path)

print("완료")
print("저장 경로:", out_path)
