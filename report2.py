from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.comments import Comment

# 입력 파일
INPUT_XLSX = Path(r"D:\Users\Qube\0. 데이터\데이터분석_최종\데이터분석_큐브_최종_장우진.xlsx")

# 출력 폴더 권한 이슈 방지용
OUTPUT_DIR = Path(r"D:\Users\장우진\dev26\qube_out_comment")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# 옵션
RENAME_SHEETS = True      # 시트 탭에 괄호 설명 붙이기
ADD_HEADER_COMMENTS = True # 헤더 셀에 주석 달기
REMOVE_DICTIONARY_SHEET = True  # data_dictionary 시트 제거

# 시트 탭에 붙일 요약 설명
sheet_desc = {
    "summary_report": "전체 데이터 요약",
    "overall_images": "전체 이미지 요약",
    "daily_questions_qd": "일자별 통계 question_data",
    "daily_questions_ql": "일자별 통계 question_list",
    "subject_stats": "과목별 통계",
    "subject_top200": "과목 상위200",
    "subject_rank_top200": "과목 랭킹 상위200",
    "subject_rank_bottom200": "과목 랭킹 하위200",
    "daily_rank_top500": "일자 랭킹 상위",
    "daily_rank_bottom500": "일자 랭킹 하위",
    "class_stats": "class별 통계",
    "class_top200": "class 상위200",
    "image_per_question_dist": "질문당 이미지 분포",
    "img_dist_with_cum": "질문당 이미지 분포 누적",
    "decoded_index": "decoded 파일 인덱스",
    "meta": "메타정보",
    "missing_before": "결측 점검 전",
    "missing_after": "결측 점검 후",
}

# 컬럼 설명
col_desc_common = {
    "rows": "해당 그룹 질문 건수",
    "valid_qid": "질문 ID QM_QST_NO 정상 건수",
    "valid_qid_rows": "질문 ID QM_QST_NO 정상 행 수",
    "date": "일자",
    "DomName": "과목 대분류",
    "SubName": "과목 소분류",
    "class_value": "반 또는 클래스 구분값",
    "image_questions": "이미지 포함 질문 건수",
    "total_images": "이미지 총 개수 합계",
    "avg_images_per_question": "질문 1건당 평균 이미지 수",
    "avg_msgs": "질문 1건당 평균 메시지 수",
    "image_rate": "이미지 포함 질문 비율 0~1",
    "image_rate_pct": "이미지 포함 질문 비율 퍼센트",
    "image_questions_rate_pct": "이미지 포함 질문 비율 퍼센트",
    "image_question_rate_pct": "이미지 포함 질문 비율 퍼센트",
    "t_first_response_s": "첫 응답까지 걸린 시간 초",
    "t_done_s": "완료까지 걸린 시간 초",
    "images_per_question": "질문당 이미지 수 0 1 2 ..",
    "questions": "해당 구간 질문 건수",
    "rate_pct": "해당 구간 비율 퍼센트",
    "cum_questions": "누적 질문 건수",
    "cum_rate_pct": "누적 비율 퍼센트",
    "generated_at": "리포트 생성 시각",
    "na_cnt": "NaN 개수",
    "na_rate_pct": "NaN 비율 퍼센트",
    "blank_like_cnt_excluding_na": "NaN 제외 빈값 유사 케이스 개수",
    "blank_like_rate_pct_excluding_na": "NaN 제외 빈값 유사 비율 퍼센트",
    "p50_first": "첫 응답 지연 t_first_response_s p50",
    "p90_first": "첫 응답 지연 t_first_response_s p90",
    "p95_first": "첫 응답 지연 t_first_response_s p95",
    "p50_done": "완료 지연 t_done_s p50",
    "p90_done": "완료 지연 t_done_s p90",
    "p95_done": "완료 지연 t_done_s p95",
}

def normalize_sheet_key(title: str) -> str:
    # 이미 괄호 설명이 붙어있으면 앞부분만 키로 사용
    t = title.strip()
    if "(" in t and t.endswith(")"):
        return t[:t.rfind("(")].strip()
    return t

def infer_col_desc(col: str) -> str:
    c = (col or "").strip()
    if not c:
        return ""

    if c.startswith(("p50_", "p90_", "p95_")):
        q = c.split("_", 1)[0]
        rest = c.split("_", 1)[1] if "_" in c else ""
        if rest == "first":
            return f"첫 응답 지연 t_first_response_s {q}"
        if rest == "done":
            return f"완료 지연 t_done_s {q}"
        return f"{rest} 지표 {q}"

    if c.endswith("_rate"):
        return "비율 0~1"
    if c.endswith("_pct"):
        return "비율 퍼센트"
    if c.endswith("_cnt"):
        return "개수"
    if c.startswith("avg_"):
        return "평균 지표"

    return ""

def safe_sheet_title(base: str, desc: str) -> str:
    base = base.strip()
    desc = (desc or "").replace(" ", "").strip()
    if not desc:
        return base[:31]

    candidate = f"{base}({desc})"
    if len(candidate) <= 31:
        return candidate

    max_desc_len = 31 - len(base) - 2
    if max_desc_len <= 0:
        return base[:31]
    return f"{base}({desc[:max_desc_len]})"

def make_unique_sheet_title(wb, desired: str) -> str:
    desired = desired[:31]
    if desired not in wb.sheetnames:
        return desired
    base = desired[:28]
    i = 2
    while True:
        cand = f"{base}_{i}"[:31]
        if cand not in wb.sheetnames:
            return cand
        i += 1

def set_comment(cell, text: str, author: str = "Qube"):
    if not text:
        return
    # 기존 주석 있으면 덮어쓰기
    cell.comment = Comment(text, author)

# 실행
wb = load_workbook(INPUT_XLSX)

# data_dictionary 제거
if REMOVE_DICTIONARY_SHEET and "data_dictionary" in wb.sheetnames:
    wb.remove(wb["data_dictionary"])

# 1) 시트명 변경 옵션
if RENAME_SHEETS:
    # 시트명 변경은 중복 충돌이 있을 수 있어 한 번에 처리
    # 우선 기존 목록 복사
    titles = [ws.title for ws in wb.worksheets]
    for t in titles:
        ws = wb[t]
        key = normalize_sheet_key(ws.title)
        desc = sheet_desc.get(key, "")
        desired = safe_sheet_title(key, desc) if desc else key[:31]
        desired = make_unique_sheet_title(wb, desired)
        ws.title = desired

# 2) 헤더 주석 달기
if ADD_HEADER_COMMENTS:
    for ws in wb.worksheets:
        # 1행을 헤더로 가정
        row1 = list(ws[1])
        if not row1:
            continue

        for cell in row1:
            val = cell.value
            if val is None:
                continue

            col = str(val).strip()
            if not col:
                continue

            desc = col_desc_common.get(col, "")
            if not desc:
                desc = infer_col_desc(col)

            if desc:
                set_comment(cell, desc, author="Qube")

# 저장
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
out_path = OUTPUT_DIR / f"{INPUT_XLSX.stem}_comment_{ts}.xlsx"
wb.save(out_path)

print("저장 완료")
print(out_path)
